# registro de vacunas 

# ideas:

# mejorar la visualizacion de los animales 
# hacer la interface 
# hacer que los mas sercanos a la fecha actual se pongan en una mini lista de 5 
# que se almacene exel si se puede (que se mantenga todo local para que sea mas seguro)

# ideas hacer que todo este en una sola paguina exepto la lista 
# que se pueda ver todod en un exel

 
import json
import os
# esto es para correrlo con un html(tambien se usa el venv)
from flask import Flask, render_template, request, redirect
#esto es para crear un exel
from openpyxl import Workbook, load_workbook
# esto es para abrir el navegadro
import webbrowser
# esto es para crear un dilay al ejecutar el ejecutable
import threading
# esto para saber la fecha actual 
from datetime import datetime, timedelta

import sys

class seguimientoDeVacunas:

    def __init__(self):
        self.listadoDeMascotas = {}

        import os

        if sys.platform == "win32":
            base_path = os.path.join(os.getenv("APPDATA"), "GestorVacunas")
        else:
            base_path = os.path.join(os.path.expanduser("~"), ".gestor_vacunas")
        data_folder = os.path.join(base_path, "data")
        os.makedirs(data_folder, exist_ok=True)

        self.data_file = os.path.join(data_folder, "registro_visible.json")
        self.exel_file = os.path.join(data_folder, "backup.xlsx")

        print("Ruta JSON:", self.data_file)

        self.cargasrTodo()
        self.arreglar_anos()
        self.guardarAimal()
    
    def guardarAimal (self):
        
        with open(self.data_file, "w") as f:
            json.dump(self.listadoDeMascotas, f, indent=4)

    def cargasrTodo (self):

        print("Ruta JSON:", self.data_file)

        if os.path.exists(self.data_file):
            with open(self.data_file, "r") as f:
                self.listadoDeMascotas = json.load(f)
        else:
            self.listadoDeMascotas = {}

    def exportar_exel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "vacunas"

        ws.append(["Nombre", "Vacuna", "Fecha"])

        for nombre, vacunas in self.listadoDeMascotas.items():
            for v in vacunas:
                ws.append([nombre, v["vacuna"], v["fecha"]])
        
        wb.save(self.exel_file)
    
    def importar_exel(self):
        if not os.path.exists(self.exel_file):
            return
        
        wb = load_workbook(self.exel_file)
        ws = wb.active

        nuevo = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            nombre, vacuna, fecha = row

            #  validacion basica
            if not nombre or not vacuna:
                continue
            
            if nombre not in nuevo:
                nuevo[nombre] = []
            
            nuevo[nombre].append({
                "vacuna": vacuna,
                "fecha": str(fecha)
            })
        self.listadoDeMascotas = nuevo
        self.guardarAimal()
    
    def convertir_fecha(self, fecha_str):
        formatos = ["%y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%Y/%m/%d"]

        for f in formatos:
            try:
                return datetime.strptime(fecha_str, f).date()
            except:
                continue
        return None
    
    def arreglar_anos(self):
        for nombre, vacunas in self.listadoDeMascotas.items():
            for v in vacunas:
                fecha = v.get("fecha", "")

                if not fecha:
                    continue
                if "/" in fecha:
                    partes = fecha.split("/")

                    if len(partes) == 3:
                        dia, mes, anio = partes

                        # si el anio tiene 2 digitos - convertir
                        if len(anio) == 2:
                            anio = "20" + anio
                        
                        # reconstruir igual formato
                        v["fecha"] = f"{dia}/{mes}/{anio}"

    def proximo_a_vencerce(self):
        hoy = datetime.now().date()
        lista = []

        for nombre, vacunas in self.listadoDeMascotas.items():
            for v in vacunas:
                
                fecha_base = self.convertir_fecha(v["fecha"])
                if not fecha_base:
                    continue

                vacuna = v["vacuna"].lower()

                # 🔥 definir frecuencia
                if "bordetela" in vacuna:
                    proxima_fecha = fecha_base + timedelta(days=180)

                elif "rabia" in vacuna:
                    proxima_fecha = fecha_base + timedelta(days=365)

                elif "quintuple" in vacuna:
                    proxima_fecha = fecha_base + timedelta(days=365)

                elif "des" in vacuna:
                    proxima_fecha = fecha_base + timedelta(days=30)

                else:
                    continue

                dias = (proxima_fecha - hoy).days

                # 🔥 SOLO próximas a vencer (7 días)
                if 0 <= dias <= 7:
                    lista.append({
                        "nombre": nombre,
                        "vacuna": v["vacuna"],
                        "fecha_aplicada": v["fecha"],
                        "proxima_fecha": proxima_fecha.strftime("%d/%m/%Y"),
                        "dias": dias
                    })

        lista.sort(key=lambda x: x["dias"])

        return lista[:5]
    
    def vacunas_vencidas(self):
        hoy = datetime.now().date()
        lista = []

        for nombre, vacunas in self.listadoDeMascotas.items():
            for v in vacunas:

                fecha_base = self.convertir_fecha(v["fecha"])
                if not fecha_base:
                    continue

                vacunas = v["vacuna"].lower()

                #mismas reglas 
                # 🔥 mismas reglas
                if "bordetela" in vacunas:
                    proxima_fecha = fecha_base + timedelta(days=180)

                elif "rabia" in vacunas:
                    proxima_fecha = fecha_base + timedelta(days=365)

                elif "quintuple" in vacunas:
                    proxima_fecha = fecha_base + timedelta(days=365)

                elif "des" in vacunas:
                    proxima_fecha = fecha_base + timedelta(days=30)

                else:
                    continue

            dias = (proxima_fecha - hoy).days

            # solo vencidas 
            if dias < 0:
                lista.append({
                    "nombre": nombre,
                    "vacuna": v["vacuna"],
                    "fecha_aplicada": v["fecha"],
                    "proxima_fecha": proxima_fecha.strftime("%d/%m/%Y"),
                    "dias": dias
                })
        
        # ordenar por mas urgente (mas negativo primero)
        lista.sort(key=lambda x: x["dias"])

        return lista[:5]

# 🔥 AQUÍ EMPIEZA FLASK

if getattr(sys, 'frozen', False):
    base_path = os.path.dirname(sys.executable)
else:
    base_path = os.path.dirname(__file__)

app = Flask(__name__, template_folder=os.path.join(base_path, "templates"))

gestor = seguimientoDeVacunas()

def abrir_navegador():
    webbrowser.open("http://127.0.0.1:5000")

# Página principal
@app.route("/")
def inicio():

    # buscador de mascotas
    busqueda = request.args.get("buscar", "").lower()
    print("Buscando: ", busqueda)

    if busqueda:
        filtrado = {
            nombre: vacunas
            for nombre, vacunas in gestor.listadoDeMascotas.items()
            if busqueda in nombre.lower()
        }
    else:
        filtrado = gestor.listadoDeMascotas
    
    proximas = gestor.proximo_a_vencerce()
    vencidas = gestor.vacunas_vencidas()
    print("proximas:", proximas)
    
    return render_template("index.html", mascotas=filtrado, proximas=proximas)


# exportar a exel
@app.route("/exportar")
def exportar():
    gestor.exportar_exel()
    return redirect("/")

#importar desde Exel
@app.route("/importar")
def importar():
    gestor.importar_exel()
    return redirect("/")

# Agregar mascota
@app.route("/agregar", methods=["POST"])
def agregar():
    nombre = request.form["nombre"]
    vacuna = request.form["vacuna"]
    fecha = request.form["fecha"]

    nueva = {"vacuna": vacuna, "fecha": fecha}

    if nombre in gestor.listadoDeMascotas:
        gestor.listadoDeMascotas[nombre].append(nueva)
    else:
        gestor.listadoDeMascotas[nombre] = [nueva]

    # gestor.listadoDeMascotas[nombre] = fecha
    gestor.guardarAimal()

    return redirect("/")

# Eliminar mascota
@app.route("/eliminar/<nombre>")
def eliminar(nombre):
    if nombre in gestor.listadoDeMascotas:
        del gestor.listadoDeMascotas[nombre]
        gestor.guardarAimal()

    return redirect("/")

# Actualizar vacuna
@app.route("/actualizar", methods=["POST"])
def actualizar():
    nombre = request.form["nombre"]
    vacunas = request.form["vacuna"]
    nueva_fecha = request.form["fecha"]

    if nombre in gestor.listadoDeMascotas:
        for v in gestor.listadoDeMascotas[nombre]:
            if v["vacuna"] == vacunas:
                v["fecha"] = nueva_fecha
                break
        gestor.guardarAimal()

    return redirect("/")

if __name__ == "__main__":
    threading.Timer(1.5, abrir_navegador).start()
    app.run(debug=True)